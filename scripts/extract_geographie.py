"""Extrait le découpage administratif officiel de la Guinée (Région > Préfecture >
Commune/Sous-préfecture > Quartier/District > Secteur) depuis les fichiers Excel du
Ministère de l'Administration du Territoire (DNAT) fournis dans Sell-It/Documents/,
vers un JSON unique bundlé dans le repo (app/data/geographie_guinee.json).

Ce script d'extraction est à usage ponctuel (source externe non versionnée, en dehors
du repo) — c'est scripts/seed_geographie.py qui charge le JSON résultant en base et qui
est le seul des deux à devoir être rejouable dans n'importe quel environnement.

Chaque région a sa propre feuille "propre" (déjà mise à plat, sans les sous-totaux
'S/T' ni les colonnes de mise en page du document officiel) :
  - Conakry (zone spéciale, pas de niveau Préfecture) : feuille 'Feuil3',
    colonnes (Communes, Quartiers/Districts, Secteurs).
  - Les 7 autres régions : feuille 'DECOUPAGE AD TERRITORIAL (2)' ou 'STATISTIC'
    selon le fichier, colonnes (Préfectures, CU/Sous-préfectures, Quartiers/Districts,
    Secteurs).
"""
import json
from pathlib import Path

import openpyxl

DOCUMENTS_DIR = Path("/Users/ibrahimadoutyoulare/Documents/Code Perso/Sell-It/Documents")
OUTPUT_PATH = Path(__file__).resolve().parents[1] / "app" / "data" / "geographie_guinee.json"

# (nom_fichier, nom_région_dans_l'appli, nom_feuille, a_un_niveau_préfecture)
SOURCES = [
    ("RA CONAKRY.xlsx", "Conakry", "Feuil3", False),
    ("RA BOKE.xlsx", "Boké", "DECOUPAGE AD TERRITORIAL (2)", True),
    ("RA FARANAH.xlsx", "Faranah", "STATISTIC", True),
    ("RA KANKAN.xlsx", "Kankan", "DECOUPAGE AD TERRITORIAL (2)", True),
    ("RA KINDIA.xlsx", "Kindia", "DECOUPAGE AD TERRITORIAL (2)", True),
    ("RA LABE.xlsx", "Labé", "DECOUPAGE AD TERRITORIAL (2)", True),
    ("RA MAMOU.xlsx", "Mamou", "DECOUPAGE AD TERRITORIAL (2)", True),
    ("RA N'ZEREKORE.xlsx", "N'Zérékoré", "STATISTIC", True),
]

HEADER_MARKERS = {"Préfectures", "Préfecture", "Communes"}


def _clean(v) -> str:
    return " ".join(str(v).strip().split()) if v is not None else ""


def extraire_region(fichier: str, feuille: str, a_prefecture: bool) -> dict:
    wb = openpyxl.load_workbook(DOCUMENTS_DIR / fichier, data_only=True)
    ws = wb[feuille]

    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        if row and row[0] in HEADER_MARKERS:
            header_row = i
            break
    if header_row is None:
        raise ValueError(f"En-tête introuvable dans {fichier} / {feuille}")

    result: dict = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if a_prefecture:
            ville, commune, quartier, secteur = (_clean(row[i]) if i < len(row) else "" for i in range(4))
        else:
            ville = None  # Conakry : pas de préfecture, la "ville" est la région elle-même
            commune, quartier, secteur = (_clean(row[i]) if i < len(row) else "" for i in range(3))
        if not commune or not quartier or not secteur:
            continue
        ville_key = ville or "__zone_speciale__"
        result.setdefault(ville_key, {}).setdefault(commune, {}).setdefault(quartier, [])
        if secteur not in result[ville_key][commune][quartier]:
            result[ville_key][commune][quartier].append(secteur)
    return result


def main() -> None:
    geographie: dict = {}
    for fichier, region, feuille, a_prefecture in SOURCES:
        data = extraire_region(fichier, feuille, a_prefecture)
        if not a_prefecture:
            # Zone spéciale de Conakry : une seule "ville" portant le nom de la région.
            data = {region: data["__zone_speciale__"]}
        geographie[region] = data

        n_villes = len(data)
        n_communes = sum(len(c) for c in data.values())
        n_quartiers = sum(len(q) for c in data.values() for q in c.values())
        n_secteurs = sum(len(s) for c in data.values() for q in c.values() for s in q.values())
        print(f"{region:14s} villes={n_villes:3d}  communes={n_communes:4d}  quartiers={n_quartiers:5d}  secteurs={n_secteurs:6d}")

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(geographie, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"\nÉcrit : {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
