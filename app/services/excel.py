"""Export comptable au format Excel réel (.xlsx, pas un CSV renommé) — CDC §3.14 :
"Export des données comptables (PDF/Excel, format compatible logiciel comptable)"."""
import io

from openpyxl import Workbook
from openpyxl.styles import Font

from app.models.schemas import CompteResultatBoutique, EcritureComptable, LigneStockValorise


def _en_tete(ws, colonnes: list[str]) -> None:
    ws.append(colonnes)
    for cell in ws[1]:
        cell.font = Font(bold=True)


def export_comptabilite_xlsx(
    comptes: list[CompteResultatBoutique],
    ecritures: list[EcritureComptable],
    stock: list[LigneStockValorise],
    noms_boutiques: dict[str, str],
) -> bytes:
    wb = Workbook()

    ws_resultat = wb.active
    ws_resultat.title = "Compte de résultat"
    _en_tete(ws_resultat, ["Boutique", "Chiffre d'affaires (GNF)", "Achats (GNF)", "Dépenses (GNF)", "Marge nette (GNF)"])
    for c in comptes:
        ws_resultat.append([
            noms_boutiques.get(c.boutique_id, c.boutique_id), c.chiffre_affaires, c.achats, c.depenses, c.marge_nette,
        ])

    ws_journal = wb.create_sheet("Journal des opérations")
    _en_tete(ws_journal, ["Date", "Boutique", "Nature", "Sens", "Montant (GNF)", "Libellé", "Auteur"])
    for e in ecritures:
        ws_journal.append([
            e.date, noms_boutiques.get(e.boutique_id, e.boutique_id), e.nature, e.sens, e.montant, e.libelle, e.auteur or "",
        ])

    ws_stock = wb.create_sheet("Stock valorisé")
    _en_tete(ws_stock, ["Boutique", "Produit", "Quantité", "Coût unitaire moyen (GNF)", "Valeur (GNF)"])
    for l in stock:
        ws_stock.append([
            noms_boutiques.get(l.boutique_id, l.boutique_id), l.produit_nom, l.quantite, l.cout_unitaire_moyen or 0, l.valeur,
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
